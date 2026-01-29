# modules/dbsync.py
from pathlib import Path
import openpyxl
from modules.db import list_all_variable_meta, get_variable_meta

INTAKE_XLSX = "intake.xlsx"
INTAKE_SHEET = "IntakeSheet"

def run_startup_sync():
    """
    Populate intake.xlsx -> IntakeSheet with all DB variables.
    Column A = Variable Group
    Column B = Variable Name
    Column C = Reserved
    Column D = Type
    Column E = Description
    All DB variables are written regardless of existing content.
    """
    intake_path = Path(INTAKE_XLSX)
    if not intake_path.exists():
        print(f"Warning: {INTAKE_XLSX} not found; skipping DB sync.")
        return

    wb = openpyxl.load_workbook(INTAKE_XLSX)
    ws = wb[INTAKE_SHEET] if INTAKE_SHEET in wb.sheetnames else wb.create_sheet(INTAKE_SHEET)

    # Clear existing content (except headers)
    for row in ws.iter_rows(min_row=2, max_col=5):
        for cell in row:
            cell.value = None

    # All DB variables
    meta_list = list_all_variable_meta()
    meta_lookup = {m["var_name"]: m for m in meta_list}

    # Group variables by category
    vars_by_category = {}
    ungrouped = []
    for meta in meta_list:
        name = meta["var_name"]
        category = meta.get("category") or "Ungrouped"
        if category == "Ungrouped":
            ungrouped.append(name)
        else:
            vars_by_category.setdefault(category, []).append(name)

    next_row = 2

    # --------------------------
    # Write grouped variables
    # --------------------------
    for category in sorted(vars_by_category.keys()):
        ws.cell(row=next_row, column=1, value=category)  # Column A = group
        next_row += 1
        for var_name in vars_by_category[category]:
            meta = meta_lookup.get(var_name) or get_variable_meta(var_name) or {}
            var_type = meta.get("var_type") or "string"
            description = meta.get("description") or ""

            ws.cell(row=next_row, column=2, value=var_name)    # Column B = variable
            ws.cell(row=next_row, column=3, value="")          # Column C = reserved
            ws.cell(row=next_row, column=4, value=var_type)    # Column D = type
            ws.cell(row=next_row, column=5, value=description) # Column E = description
            next_row += 1

    # --------------------------
    # Write ungrouped variables
    # --------------------------
    if ungrouped:
        ws.cell(row=next_row, column=1, value="Ungrouped Variables")
        next_row += 1
        for var_name in ungrouped:
            meta = meta_lookup.get(var_name) or get_variable_meta(var_name) or {}
            var_type = meta.get("var_type") or "string"
            description = meta.get("description") or ""

            ws.cell(row=next_row, column=2, value=var_name)
            ws.cell(row=next_row, column=3, value="")
            ws.cell(row=next_row, column=4, value=var_type)
            ws.cell(row=next_row, column=5, value=description)
            next_row += 1

    wb.save(INTAKE_XLSX)
    print("DB sync complete. Variables written with grouping, types in D, descriptions in E.")
