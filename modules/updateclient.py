# modules/updateclient.py
import tkinter as tk
from tkinter import messagebox, font
import openpyxl

from modules.db import (
    list_clients,
    set_variable,
    get_variable_meta,
    set_variable_meta,
    delete_client,
    get_all_variables_for_client,
    list_all_variable_meta,
    get_variables,
)

INTAKE_XLSX = "intake.xlsx"
INTAKE_SHEET = "IntakeSheet"


# -------------------------------------------------
# Client display helper
# -------------------------------------------------
def build_client_label(client_id):
    vars_ = get_variables("client", client_id)
    matterid = vars_.get("matterid", "")
    first = vars_.get("firstname", "")
    last = vars_.get("lastname", "")

    name = " ".join(p for p in [first, last] if p).strip()
    label = f"ID-{client_id}"
    if matterid:
        label = f"({matterid}) - {label}"
    if name:
        label += f" - {name}"
    return label


# -------------------------------------------------
# Client selection popup
# -------------------------------------------------
def select_client(clients):
    selected_id = tk.IntVar(value=-1)
    cancelled = tk.BooleanVar(value=False)

    popup = tk.Toplevel()
    popup.title("Select Client")
    popup.geometry("450x450")
    popup.grab_set()

    tk.Label(popup, text="Search Client:").pack(pady=(10, 0))
    search_var = tk.StringVar()

    list_frame = tk.Frame(popup)
    list_frame.pack(fill="both", expand=True, padx=10)

    client_ids = [c[0] for c in clients]

    def update_list(*_):
        term = search_var.get().lower().strip()
        for w in list_frame.winfo_children():
            w.destroy()

        matches = []
        for cid in client_ids:
            label = build_client_label(cid)
            if not term or term in label.lower():
                matches.append((cid, label))

        if not matches:
            tk.Label(list_frame, text="No matches found").pack(anchor="w")
            return

        for cid, label in matches:
            tk.Radiobutton(
                list_frame,
                text=label,
                variable=selected_id,
                value=cid,
                wraplength=380,
            ).pack(anchor="w")

    search_var.trace_add("write", update_list)

    tk.Entry(popup, textvariable=search_var, width=50).pack(pady=5)
    update_list()
    tk.Entry(popup, textvariable=search_var, width=50).focus_set()

    def submit():
        if selected_id.get() == -1:
            messagebox.showwarning("Select Client", "No client selected.", parent=popup)
            return
        popup.destroy()

    def cancel():
        cancelled.set(True)
        popup.destroy()

    btns = tk.Frame(popup)
    btns.pack(pady=10)
    tk.Button(btns, text="Select", command=submit).pack(side="left", padx=5)
    tk.Button(btns, text="Cancel", command=cancel).pack(side="left", padx=5)

    popup.wait_window()

    if cancelled.get():
        return None, True
    return selected_id.get(), False


# -------------------------------------------------
# Intake sheet variable loader
# -------------------------------------------------
def load_intake_variables():
    vars_from_excel = set()
    try:
        wb = openpyxl.load_workbook(INTAKE_XLSX, data_only=True)
        if INTAKE_SHEET not in wb.sheetnames:
            return vars_from_excel

        sheet = wb[INTAKE_SHEET]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            var_name = row[1]  # Column B
            if var_name and isinstance(var_name, str):
                vars_from_excel.add(var_name.strip())
    except Exception:
        pass
    return vars_from_excel


# -------------------------------------------------
# Main update workflow
# -------------------------------------------------
def update_client():
    while True:
        clients = list_clients()
        if not clients:
            messagebox.showinfo("Update Client", "No clients found.")
            return

        client_id, cancelled = select_client(clients)
        if cancelled or client_id is None:
            return

        # Load all base variables for this client
        client_vars = get_all_variables_for_client("client", client_id)

        # --- Compute derived variables automatically ---
        for var_meta in list_all_variable_meta():
            if var_meta.get("is_derived"):
                expr = var_meta.get("derived_expression")
                try:
                    client_vars[var_meta["var_name"]] = eval(expr, {}, client_vars)
                except Exception:
                    client_vars[var_meta["var_name"]] = ""  # fallback if evaluation fails

        window = tk.Toplevel()
        window.title(f"Update Client Variables — {build_client_label(client_id)}")
        window.geometry("980x620")
        window.grab_set()

        tk.Label(
            window,
            text="Edit variables or add new ones. Derived variables are computed automatically.",
            font=("Helvetica", 13),
        ).pack(pady=8)

        # --- Search bar ---
        tk.Label(window, text="Search Variable:").pack()
        var_search = tk.StringVar()
        tk.Entry(window, textvariable=var_search, width=50).pack(pady=(0, 10))


        # --- Attorney Assignment Section ---
        attorney_frame = tk.Frame(window)
        attorney_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Label(attorney_frame, text="Opposing Counsel:", font=("Helvetica", 11, "bold")).pack(side="left", padx=5)
        
        # Get current attorney assignment DIRECTLY from clients table
        from modules.db import list_opposing_counsel
        import sqlite3
        from modules.db import DB_PATH
        
        # Read opposing_counsel_id directly from clients table
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT opposing_counsel_id FROM clients WHERE id=?", (client_id,))
        row = c.fetchone()
        conn.close()
        
        current_counsel_id = row[0] if row and row[0] else None
        
        attorneys = list_opposing_counsel()
        attorney_options = ["(None)"] + [f"{row[1]} {row[2]} - {row[11] or 'No Firm'}" for row in attorneys]
        attorney_ids = [None] + [row[0] for row in attorneys]
        
        selected_attorney_idx = 0
        if current_counsel_id:
            try:
                selected_attorney_idx = attorney_ids.index(int(current_counsel_id))
            except (ValueError, TypeError):
                pass
        
        attorney_var = tk.StringVar(value=attorney_options[selected_attorney_idx])
        attorney_dropdown = tk.OptionMenu(attorney_frame, attorney_var, *attorney_options)
        attorney_dropdown.config(width=40)
        attorney_dropdown.pack(side="left", padx=5)
        
        def save_attorney_assignment():
            selected_name = attorney_var.get()
            selected_idx = attorney_options.index(selected_name)
            selected_id = attorney_ids[selected_idx]
            
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            
            if selected_id:
                c.execute("UPDATE clients SET opposing_counsel_id = ? WHERE id = ?", (selected_id, client_id))
                conn.commit()
                conn.close()
                messagebox.showinfo("Success", "Opposing counsel assigned!", parent=window)
            else:
                c.execute("UPDATE clients SET opposing_counsel_id = NULL WHERE id = ?", (client_id,))
                conn.commit()
                conn.close()
                messagebox.showinfo("Success", "Opposing counsel removed!", parent=window)
        
        tk.Button(attorney_frame, text="Assign Attorney", command=save_attorney_assignment).pack(side="left", padx=5)




        # --- Scrollable table setup ---
        container = tk.Frame(window)
        container.pack(fill="both", expand=True, padx=10, pady=5)

        table_canvas = tk.Canvas(container)
        table_frame = tk.Frame(table_canvas)
        scrollbar = tk.Scrollbar(container, orient="vertical", command=table_canvas.yview)
        table_canvas.configure(yscrollcommand=scrollbar.set)

        table_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        table_canvas.create_window((0, 0), window=table_frame, anchor="nw")

        def resize_canvas(event):
            table_canvas.configure(scrollregion=table_canvas.bbox("all"))

        table_frame.bind("<Configure>", resize_canvas)

        # --- Table headers ---
        headers = ["Variable", "Description", "Value"]
        widths = [25, 45, 25]
        for i, (h, w) in enumerate(zip(headers, widths)):
            tk.Label(
                table_frame,
                text=h,
                width=w,
                font=("Helvetica", 10, "bold"),
                anchor="w",
            ).grid(row=0, column=i, sticky="w")

        all_var_meta = list_all_variable_meta()
        meta_lookup = {m["var_name"]: m for m in all_var_meta}
        all_var_names = set(meta_lookup.keys()) | set(client_vars.keys())

        row_vars = {}
        for row_index, var_name in enumerate(sorted(all_var_names), start=1):
            meta = meta_lookup.get(var_name) or {}
            raw_val = client_vars.get(var_name, "")
            value = raw_val.get("value") if isinstance(raw_val, dict) else raw_val or ""

            val_var = tk.StringVar(value=value)
            desc_var = tk.StringVar(value=meta.get("description", ""))

            bg_color = "#e6f7ff" if meta.get("is_derived") else None

            tk.Label(table_frame, text=var_name, width=25, anchor="w", bg=bg_color).grid(row=row_index, column=0, sticky="w")
            tk.Entry(table_frame, textvariable=desc_var, width=45, state="readonly", bg=bg_color).grid(row=row_index, column=1, sticky="w")
            tk.Entry(table_frame, textvariable=val_var, width=25, bg=bg_color).grid(row=row_index, column=2, sticky="w")

            row_vars[var_name] = {"val": val_var, "desc": desc_var, "meta": meta}

        # --- Apply updates function ---
        def apply_updates():
            changed_vars = []
            for var, info in row_vars.items():
                new_val = info["val"].get().strip()
                meta = info["meta"]
                old_val = client_vars.get(var, "")
                if isinstance(old_val, dict):
                    old_val = old_val.get("value", "")
                old_val = old_val or ""
                if new_val != old_val:
                    changed_vars.append((var, meta.get("description", ""), new_val, old_val))

            if not changed_vars:
                messagebox.showinfo("Update Client", "No changes detected.")
                return

            confirm_win = tk.Toplevel()
            confirm_win.title("Confirm Changes")
            confirm_win.geometry("800x400")
            confirm_win.grab_set()
            tk.Label(
                confirm_win,
                text="You are adding/changing values for these variables:",
                font=("Helvetica", 12, "bold"),
            ).pack(pady=5)

            list_frame = tk.Frame(confirm_win)
            list_frame.pack(fill="both", expand=True, padx=10, pady=5)
            headers = ["Variable", "Description", "New Value", "Old Value"]
            for col, h in enumerate(headers):
                tk.Label(list_frame, text=h, font=("Helvetica", 10, "bold")).grid(row=0, column=col, sticky="w", padx=2)

            for i, (var, desc, new_val, old_val) in enumerate(changed_vars, start=1):
                tk.Label(list_frame, text=var, anchor="w").grid(row=i, column=0, sticky="w", padx=2)
                tk.Label(list_frame, text=desc, anchor="w").grid(row=i, column=1, sticky="w", padx=2)
                tk.Label(list_frame, text=new_val, anchor="w", bg="#ffff99").grid(row=i, column=2, sticky="w", padx=2)
                tk.Label(list_frame, text=old_val, anchor="w").grid(row=i, column=3, sticky="w", padx=2)

            submitted = tk.BooleanVar(value=False)

            def confirm():
                for var, _, val, _ in changed_vars:
                    set_variable("client", client_id, var, val)
                submitted.set(True)
                confirm_win.destroy()
                window.destroy()

            def cancel():
                submitted.set(True)
                confirm_win.destroy()

            btn_frame = tk.Frame(confirm_win)
            btn_frame.pack(pady=10)
            tk.Button(btn_frame, text="Yes", command=confirm).pack(side="left", padx=5)
            tk.Button(btn_frame, text="No", command=cancel).pack(side="left", padx=5)

            confirm_win.wait_variable(submitted)

        # --- Bottom buttons in fixed frame ---
        btns = tk.Frame(window)
        btns.pack(side="bottom", pady=10)
        tk.Button(btns, text="Apply Updates", command=apply_updates).pack(side="left", padx=5)
        tk.Button(btns, text="Delete Client", command=lambda: delete_client_workflow(client_id, window), fg="red").pack(side="left", padx=5)
        tk.Button(btns, text="Cancel", command=window.destroy).pack(side="left", padx=5)


        window.wait_window()


def delete_client_workflow(client_id, window):
    if not messagebox.askyesno(
        "Delete Client",
        "⚠ WARNING ⚠\n\nThis will permanently delete this client.",
        parent=window,
    ):
        return
    delete_client(client_id)
    window.destroy()
    messagebox.showinfo("Delete Client", "Client deleted.")
