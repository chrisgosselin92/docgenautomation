# modules/intake.py
from pathlib import Path
import pandas as pd
from tkinter import messagebox
import subprocess
import platform

from modules.db import (
    create_client,
    set_variable,
    set_variable_meta,
    variable_exists,
    list_clients,
    get_variable_meta,
    get_variables,
)

import openpyxl

INTAKE_FILE = Path("intake.xlsx")
INTAKE_SHEET = "IntakeSheet"

SUPPORTED_TYPES = {"string", "int", "float", "bool", "date", "iso-date"}


# ---------------------------
# Helpers
# ---------------------------
def normalize_type(raw_type):
    if raw_type is None or pd.isna(raw_type):
        return "string"
    t = str(raw_type).strip().lower()
    return t if t in SUPPORTED_TYPES else "string"


def coerce_value(value, var_type):
    if value is None or pd.isna(value) or value == "":
        return None
    if isinstance(value, (dict, list, tuple)):
        return None
    try:
        if var_type == "int":
            return int(value)
        if var_type == "float":
            return float(value)
        if var_type == "bool":
            return str(value).strip().lower() in {"1", "true", "yes", "y"}
        return str(value).strip()
    except Exception:
        return None


def open_intake_file():
    if platform.system() == "Windows":
        subprocess.run(["start", str(INTAKE_FILE)], shell=True)
    elif platform.system() == "Darwin":
        subprocess.run(["open", str(INTAKE_FILE)])
    else:
        subprocess.run(["xdg-open", str(INTAKE_FILE)])


# ---------------------------
# Intake import
# ---------------------------
def import_intake_for_client(client_id=None):
    if not INTAKE_FILE.exists():
        messagebox.showerror("Missing File", "intake.xlsx not found.")
        return

    try:
        df = pd.read_excel(INTAKE_FILE, sheet_name=INTAKE_SHEET)
    except Exception:
        messagebox.showerror("Missing Sheet", f"Sheet '{INTAKE_SHEET}' not found.")
        return

    if df.empty or len(df) < 2:
        messagebox.showwarning("Empty Intake", "Intake sheet contains no usable rows.")
        return

    # Determine client_id if not provided
    if client_id is None:
        matterid_row = df[df.iloc[:, 1].astype(str).str.strip().str.lower() == "matterid"]
        matterid = None
        if not matterid_row.empty:
            first_val = matterid_row.iloc[0, 2]  # Column C has the matterid value
            matterid = str(first_val).strip() if pd.notna(first_val) else None

        if not matterid:
            messagebox.showwarning(
                "Import Cancelled",
                "No matterid found in the intake sheet. Please fill it in column C next to 'matterid'."
            )
            open_intake_file()
            return

        existing_clients = list_clients()
        client_lookup = {c[4]: c[0] for c in existing_clients}
        if matterid in client_lookup:
            messagebox.showwarning(
                "Client Exists",
                f"A client with matterid '{matterid}' already exists in the database."
            )
            return
        else:
            client_id = create_client(matterid)

    # Fetch all DB variables for the client
    client_vars = get_variables("client", client_id)

    # Load workbook directly with openpyxl
    wb = openpyxl.load_workbook(INTAKE_FILE)
    ws = wb[INTAKE_SHEET]

    updates = 0
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=ws.max_column), start=2):
        var_cell = row[1]  # Column B: variable name
        val_cell = row[2]  # Column C: value
        type_cell = row[3] if len(row) > 3 else None  # Column D: type
        desc_cell = row[4] if len(row) > 4 else None  # Column E: description

        var_name = str(var_cell.value).strip() if var_cell.value else None
        if not var_name:
            val_cell.value = None
            continue

        # Determine type and description
        var_type = normalize_type(type_cell.value if type_cell else None)
        description = str(desc_cell.value).strip() if desc_cell and desc_cell.value else ""

        # Ensure metadata exists or update description
        meta = get_variable_meta(var_name)
        if meta:
            if description and description != (meta.get("description") or ""):
                set_variable_meta(
                    var_name=var_name,
                    var_type=meta.get("var_type", "string"),
                    description=description,
                    category=meta.get("category", "General"),
                    display_order=meta.get("display_order", 0),
                    is_derived=meta.get("is_derived", 0),
                    derived_expression=meta.get("derived_expression")
                )
        else:
            set_variable_meta(var_name=var_name, var_type=var_type, description=description)

        # Read value from Excel and coerce
        raw_value = val_cell.value
        value = coerce_value(raw_value, var_type)
        if value is not None:
            set_variable("client", client_id, var_name, value)
            updates += 1

        # Clear Excel column C except header
        val_cell.value = None if r_idx > 1 else val_cell.value

    wb.save(INTAKE_FILE)

    messagebox.showinfo(
        "Intake Imported",
        f"{updates} value(s) imported for client ID {client_id}."
    )
